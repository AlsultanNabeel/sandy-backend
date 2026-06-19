"""Extract plain text from uploaded Telegram documents.

Supported: TXT, MD, PDF, DOCX, CSV, JSON, XLSX, and source-code files.
Returns (text, error_message) — error_message is None on success.
"""

from __future__ import annotations

import csv
import io
import json
import os

_MAX_CHARS = 12_000  # اقتصاد tokens — نقطع بعدها


def _truncate(text: str, label: str = "") -> str:
    if len(text) <= _MAX_CHARS:
        return text.strip()
    note = f"\n\n… [مقتطع — {label}الملف أطول من {_MAX_CHARS} حرف]"
    return text[:_MAX_CHARS].strip() + note


# extractors

def _read_plain(data: bytes) -> tuple[str, str | None]:
    for enc in ("utf-8", "utf-16", "cp1256", "latin-1"):
        try:
            return data.decode(enc), None
        except (UnicodeDecodeError, LookupError):
            continue
    return "", "تعذّر قراءة ترميز الملف."


def _read_pdf(data: bytes) -> tuple[str, str | None]:
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(data))
        parts = []
        for page in reader.pages:
            t = page.extract_text() or ""
            if t.strip():
                parts.append(t)
        text = "\n\n".join(parts)
        if not text.strip():
            return "", "الـ PDF لا يحتوي نصاً قابلاً للقراءة (قد يكون صور فقط)."
        return text, None
    except Exception as exc:
        return "", f"خطأ في قراءة PDF: {exc}"


def _read_docx(data: bytes) -> tuple[str, str | None]:
    try:
        from docx import Document
        doc = Document(io.BytesIO(data))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paragraphs), None
    except Exception as exc:
        return "", f"خطأ في قراءة DOCX: {exc}"


def _read_csv(data: bytes) -> tuple[str, str | None]:
    text, err = _read_plain(data)
    if err:
        return "", err
    try:
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return "", "الـ CSV فارغ."
        col_widths = [max(len(str(row[i])) if i < len(row) else 0 for row in rows) for i in range(len(rows[0]))]
        lines = []
        for row in rows:
            lines.append(" | ".join(str(cell).ljust(col_widths[i]) for i, cell in enumerate(row)))
        return "\n".join(lines), None
    except Exception:
        return text, None


def _read_xlsx(data: bytes) -> tuple[str, str | None]:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        parts = []
        for sheet in wb.worksheets:
            rows_text = []
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    rows_text.append(" | ".join(cells))
            if rows_text:
                parts.append(f"### {sheet.title}\n" + "\n".join(rows_text))
        return "\n\n".join(parts) or "", None
    except Exception as exc:
        return "", f"خطأ في قراءة XLSX: {exc}"


def _read_json(data: bytes) -> tuple[str, str | None]:
    text, err = _read_plain(data)
    if err:
        return "", err
    try:
        parsed = json.loads(text)
        return json.dumps(parsed, ensure_ascii=False, indent=2), None
    except Exception:
        return text, None


# extension map

_TEXT_EXTENSIONS = {
    ".txt", ".md", ".markdown", ".rst", ".log", ".yaml", ".yml",
    ".toml", ".ini", ".cfg", ".env",
    ".py", ".js", ".ts", ".jsx", ".tsx", ".java", ".c", ".cpp",
    ".h", ".cs", ".go", ".rs", ".rb", ".php", ".swift", ".kt",
    ".html", ".htm", ".xml", ".css", ".scss", ".sql",
    ".sh", ".bash", ".zsh", ".bat", ".ps1",
}


def extract_text(file_bytes: bytes, filename: str) -> tuple[str, str | None]:
    """Return (text, error). error is None on success."""
    ext = os.path.splitext(filename.lower())[1]

    if ext == ".pdf":
        text, err = _read_pdf(file_bytes)
    elif ext == ".doc":
        # python-docx can't read the legacy binary .doc format — only .docx.
        return "", "صيغة .doc القديمة غير مدعومة — احفظ الملف كـ .docx وأعد إرساله."
    elif ext == ".docx":
        text, err = _read_docx(file_bytes)
    elif ext == ".csv":
        text, err = _read_csv(file_bytes)
    elif ext in (".xlsx", ".xls"):
        text, err = _read_xlsx(file_bytes)
    elif ext == ".json":
        text, err = _read_json(file_bytes)
    elif ext in _TEXT_EXTENSIONS:
        text, err = _read_plain(file_bytes)
    else:
        # محاولة أخيرة كـ plain text
        text, err = _read_plain(file_bytes)
        if err:
            return "", f"صيغة الملف ({ext or 'غير معروفة'}) غير مدعومة حالياً."

    if err:
        return "", err
    return _truncate(text, f"{filename} — "), None
